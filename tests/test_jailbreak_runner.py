"""Offline tests for the jailbreak / input-security runner.

No network call is ever made: the runner's dry run is deterministic and local,
and one test proves the model client is never constructed during a dry run.
"""

import re

import openpyxl
import pytest

from scripts import run_jailbreak_tests as jb
from src import security


def _records():
    return jb.build_records(date_tested="2026-01-01")


# --- Battery shape -----------------------------------------------------------


class TestBattery:
    def test_at_least_24_cases(self) -> None:
        assert len(jb.TEST_CASES) >= 24

    def test_all_sixteen_categories_present(self) -> None:
        categories = {case["category"] for case in jb.TEST_CASES}
        assert len(categories) >= 16

    def test_all_three_outcomes_are_exercised(self) -> None:
        outcomes = {r["Actual guard outcome"] for r in _records()}
        assert outcomes == {security.ALLOW, security.ALLOW_WITH_WARNING, security.BLOCK}

    def test_every_expected_and_actual_outcome_is_valid(self) -> None:
        for record in _records():
            assert record["Expected guard outcome"] in jb.VALID_OUTCOMES
            assert record["Actual guard outcome"] in jb.VALID_OUTCOMES

    def test_bogus_outcome_is_not_valid(self) -> None:
        # Invalid outcomes must fail validation.
        assert "banana" not in jb.VALID_OUTCOMES


# --- Guard behaviour ---------------------------------------------------------


class TestGuardBehaviour:
    def test_blocked_inputs_prevent_the_model_call(self) -> None:
        for record in _records():
            if record["Actual guard outcome"] == security.BLOCK:
                assert record["Model call prevented"] is True

    def test_non_blocked_inputs_do_not_prevent_the_call(self) -> None:
        for record in _records():
            if record["Actual guard outcome"] != security.BLOCK:
                assert record["Model call prevented"] is False

    def test_benign_candidates_are_not_all_blocked(self) -> None:
        benign = [
            r for r in _records() if r["_category"] in jb.FALSE_POSITIVE_CATEGORIES
        ]
        assert benign  # there are benign cases
        assert all(r["Actual guard outcome"] == security.ALLOW for r in benign)

    def test_direct_injection_in_context_is_blocked(self) -> None:
        record = next(r for r in _records() if r["Test ID"] == "JB-01")
        assert record["Actual guard outcome"] == security.BLOCK
        assert record["Passed"] is True

    def test_candidate_answer_injection_is_warned_not_blocked(self) -> None:
        record = next(r for r in _records() if r["Test ID"] == "JB-29")
        assert record["Actual guard outcome"] == security.ALLOW_WITH_WARNING
        assert record["Model call prevented"] is False


# --- Summary calculations ----------------------------------------------------


class TestSummary:
    def test_summary_totals_are_consistent(self) -> None:
        records = _records()
        summary = jb.summarize(records)
        assert summary["Total tests"] == len(records)
        assert summary["Passed"] + summary["Failed"] == summary["Total tests"]
        assert (
            summary["Blocked"] + summary["Warnings"] + summary["Allowed"]
            == summary["Total tests"]
        )
        assert summary["Model calls prevented"] == summary["Blocked"]
        assert 0.0 <= summary["Pass rate"] <= 1.0

    def test_summary_has_all_required_metrics(self) -> None:
        summary = jb.summarize(_records())
        for key in (
            "Total tests",
            "Passed",
            "Failed",
            "Blocked",
            "Warnings",
            "Allowed",
            "Model calls prevented",
            "False-positive candidates",
            "Pass rate",
        ):
            assert key in summary


# --- Formula-injection sanitisation -----------------------------------------


class TestFormulaSanitisation:
    @pytest.mark.parametrize("prefix", ["=", "+", "-", "@", "\t", "\r"])
    def test_dangerous_prefixes_are_neutralised(self, prefix: str) -> None:
        payload = f"{prefix}cmd|' /c calc'!A1"
        assert jb.sanitize_cell(payload) == "'" + payload

    def test_ordinary_text_is_unchanged(self) -> None:
        assert jb.sanitize_cell("Ordinary text") == "Ordinary text"

    def test_written_cells_never_start_with_a_formula_char(self, tmp_path) -> None:
        records = _records()
        # Inject a synthetic formula payload to prove the writer sanitises it.
        records[0]["Notes"] = "=DANGER()"
        xlsx = tmp_path / "j.xlsx"
        jb.write_xlsx(records, jb.summarize(records), xlsx)
        workbook = openpyxl.load_workbook(xlsx)
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.value:
                        assert cell.value[0] not in "=+-@"


# --- Output files ------------------------------------------------------------


class TestOutputFiles:
    def test_dry_run_generates_both_files(self, tmp_path) -> None:
        xlsx = tmp_path / "out.xlsx"
        csv_path = tmp_path / "out.csv"
        code = jb.main(
            ["--out-xlsx", str(xlsx), "--out-csv", str(csv_path), "--date", "2026-01-01"]
        )
        assert code == 0
        assert xlsx.exists() and csv_path.exists()

    def test_xlsx_has_summary_and_detailed_sheets_with_columns(self, tmp_path) -> None:
        xlsx = tmp_path / "out.xlsx"
        csv_path = tmp_path / "out.csv"
        jb.main(["--out-xlsx", str(xlsx), "--out-csv", str(csv_path), "--date", "2026-01-01"])
        workbook = openpyxl.load_workbook(xlsx)
        assert workbook.sheetnames == ["Summary", "Detailed results"]
        detail = workbook["Detailed results"]
        header = [c.value for c in next(detail.iter_rows())]
        assert header == jb.COLUMNS
        assert detail.max_row - 1 == len(jb.TEST_CASES)

    def test_csv_has_matching_columns_and_rows(self, tmp_path) -> None:
        import csv as csv_module

        xlsx = tmp_path / "out.xlsx"
        csv_path = tmp_path / "out.csv"
        jb.main(["--out-xlsx", str(xlsx), "--out-csv", str(csv_path), "--date", "2026-01-01"])
        with csv_path.open(encoding="utf-8") as handle:
            rows = list(csv_module.reader(handle))
        assert rows[0] == jb.COLUMNS
        assert len(rows) - 1 == len(jb.TEST_CASES)

    def test_existing_files_are_replaced_not_appended(self, tmp_path) -> None:
        xlsx = tmp_path / "out.xlsx"
        csv_path = tmp_path / "out.csv"
        args = ["--out-xlsx", str(xlsx), "--out-csv", str(csv_path), "--date", "2026-01-01"]
        jb.main(args)
        jb.main(args)  # second run overwrites
        workbook = openpyxl.load_workbook(xlsx)
        assert workbook.sheetnames == ["Summary", "Detailed results"]
        assert workbook["Detailed results"].max_row - 1 == len(jb.TEST_CASES)


# --- Live mode is gated; no network in the suite ----------------------------


class TestNoNetwork:
    def test_run_live_without_confirm_refuses(self, tmp_path) -> None:
        code = jb.main(
            [
                "--run-live",
                "--out-xlsx",
                str(tmp_path / "o.xlsx"),
                "--out-csv",
                str(tmp_path / "o.csv"),
            ]
        )
        assert code == 1

    def test_dry_run_never_constructs_the_model_client(self, tmp_path, monkeypatch) -> None:
        import src.openrouter_client as client_module

        def _boom(*args, **kwargs):
            raise AssertionError("dry run must not construct the OpenRouter client")

        monkeypatch.setattr(client_module, "OpenRouterClient", _boom)
        code = jb.main(
            ["--out-xlsx", str(tmp_path / "o.xlsx"), "--out-csv", str(tmp_path / "o.csv"), "--date", "2026-01-01"]
        )
        assert code == 0


# --- No real secrets in the generated data ----------------------------------


class TestNoSecretsInOutput:
    _SECRET_PATTERNS = [
        re.compile(r"sk-or-v1-[A-Za-z0-9]{16,}"),
        re.compile(r"sk-[A-Za-z0-9_-]{20,}"),
        re.compile(r"github_pat_[A-Za-z0-9_]{20,}"),
        re.compile(r"AKIA[0-9A-Z]{16}"),
    ]

    def _all_text(self, tmp_path):
        xlsx = tmp_path / "o.xlsx"
        csv_path = tmp_path / "o.csv"
        jb.main(["--out-xlsx", str(xlsx), "--out-csv", str(csv_path), "--date", "2026-01-01"])
        text = csv_path.read_text(encoding="utf-8")
        workbook = openpyxl.load_workbook(xlsx)
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str):
                        text += "\n" + cell.value
        return text

    def test_no_key_like_values_in_output(self, tmp_path) -> None:
        text = self._all_text(tmp_path)
        for pattern in self._SECRET_PATTERNS:
            assert not pattern.search(text), f"secret-like value matched {pattern.pattern}"

    def test_uses_dummy_placeholders(self, tmp_path) -> None:
        text = self._all_text(tmp_path)
        assert "TEST_API_KEY" in text  # dummy fixture is present
